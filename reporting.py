#!/usr/bin/env python3
"""
Raporlama Sistemi
Tahmin loglarını tutar ve günlük Excel raporları oluşturur.
"""

import csv
import os
import pandas as pd
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Optional, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from constants import ReportConstants


def ensure_dirs():
    """logs/ ve reports/ klasörlerini oluştur"""
    dirs = ['logs', 'reports']
    for dir_name in dirs:
        Path(dir_name).mkdir(exist_ok=True)
        print(f"📁 Klasör hazır: {dir_name}/")


def append_prediction_log(row: Dict[str, Union[str, float]], path: str = "logs/predictions.csv"):
    """
    Tahmin loguna satır ekle
    
    Args:
        row (dict): Tahmin verisi
            - timestamp: str
            - sicaklik: float
            - titresim: float  
            - tork: float
            - rul: float
            - status: str
        path (str): Log dosyası yolu
    """
    try:
        # Klasörü oluştur
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        
        # Gerekli alanları kontrol et
        required_fields = ['timestamp', 'sicaklik', 'titresim', 'tork', 'rul', 'status']
        for field in required_fields:
            if field not in row:
                raise ValueError(f"Gerekli alan eksik: {field}")
        
        # Dosya var mı kontrol et
        file_exists = Path(path).exists()
        
        # CSV'ye yaz
        with open(path, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=required_fields)
            
            # Başlık yaz (gerekirse)
            if not file_exists:
                writer.writeheader()
                print(f"📝 Yeni log dosyası oluşturuldu: {path}")
            
            # Veriyi yaz
            writer.writerow(row)
            
    except OSError as e:
        raise OSError(f"Log yazma hatası: {str(e)}") from e


def daily_report_to_excel(date_str: Optional[str] = None) -> str:
    """
    Günlük Excel raporu oluştur
    
    Args:
        date_str (str, optional): YYYY-MM-DD formatında tarih. None ise bugün.
    
    Returns:
        str: Oluşturulan Excel dosyasının yolu
    
    Raises:
        Exception: Dosya bulunamadığında veya veri işleme hatalarında
    """
    try:
        # Tarih belirle
        if date_str is None:
            target_date = date.today()
            date_str = target_date.strftime('%Y-%m-%d')
        else:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                raise ValueError(f"Geçersiz tarih formatı: {date_str}. YYYY-MM-DD formatında olmalı.")
        
        # Log dosyasını kontrol et
        log_path = "logs/predictions.csv"
        if not Path(log_path).exists():
            raise FileNotFoundError(f"Log dosyası bulunamadı: {log_path}")
        
        # CSV'yi oku
        df = pd.read_csv(log_path)
        
        if df.empty:
            raise ValueError("Log dosyası boş")
        
        # Timestamp sütununu datetime'a çevir
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        df['date'] = df['timestamp'].dt.date
        
        # İlgili tarihteki verileri filtrele
        daily_data = df[df['date'] == target_date].copy()
        
        if daily_data.empty:
            raise ValueError(f"{date_str} tarihinde veri bulunamadı")
        
        # Rapor klasörünü oluştur
        ensure_dirs()
        
        # Excel dosya yolu
        excel_path = f"reports/report_{date_str}.xlsx"
        
        # Excel dosyası oluştur
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            
            # Sheet1: Ham kayıtlar
            daily_data_clean = daily_data.drop('date', axis=1)  # date sütununu kaldır
            daily_data_clean.to_excel(writer, sheet_name=ReportConstants.RAW_DATA_SHEET, index=False)
            
            # Sheet2: Özet
            summary_data = {
                'Metrik': [
                    'Toplam Kayıt Sayısı',
                    'CRITICAL Sayısı',
                    'PLANNED Sayısı', 
                    'NORMAL Sayısı',
                    'UNKNOWN Sayısı',
                    'Ortalama RUL',
                    'Maksimum RUL',
                    'Minimum RUL'
                ],
                'Değer': [
                    len(daily_data),
                    len(daily_data[daily_data['status'] == 'CRITICAL']),
                    len(daily_data[daily_data['status'] == 'PLANNED']),
                    len(daily_data[daily_data['status'] == 'NORMAL']),
                    len(daily_data[daily_data['status'] == 'UNKNOWN']),
                    round(daily_data['rul'].mean(), 2),
                    round(daily_data['rul'].max(), 2),
                    round(daily_data['rul'].min(), 2)
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Özet', index=False)
        
        # Excel dosyasını biçimlendir
        _format_excel_report(excel_path)
        
        print(f"📊 Günlük rapor oluşturuldu: {excel_path}")
        print(f"📅 Tarih: {date_str}")
        print(f"📈 Toplam kayıt: {len(daily_data)}")
        
        return excel_path
        
    except (OSError, ValueError) as e:
        raise OSError(f"Excel raporu oluşturma hatası: {str(e)}") from e


def _format_excel_report(excel_path: str):
    """Excel raporunu biçimlendir"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        # Sheet1 biçimlendirme
        if ReportConstants.RAW_DATA_SHEET in wb.sheetnames:
            ws1 = wb[ReportConstants.RAW_DATA_SHEET]
            
            # Başlık satırını biçimlendir
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws1[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Sütun genişliklerini ayarla
            ws1.column_dimensions['A'].width = 20  # timestamp
            ws1.column_dimensions['B'].width = 12  # sicaklik
            ws1.column_dimensions['C'].width = 12  # titresim
            ws1.column_dimensions['D'].width = 12  # tork
            ws1.column_dimensions['E'].width = 12  # rul
            ws1.column_dimensions['F'].width = 15  # status
        
        # Sheet2 biçimlendirme
        if 'Özet' in wb.sheetnames:
            ws2 = wb['Özet']
            
            # Başlık satırını biçimlendir
            for cell in ws2[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Sütun genişliklerini ayarla
            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 15
        
        wb.save(excel_path)
        
    except OSError as e:
        print(f"⚠️  Excel biçimlendirme hatası: {e}")


def test_reporting_system():
    """Raporlama sistemini test et"""
    print("🧪 Raporlama sistemi test ediliyor...")
    
    # Klasörleri oluştur
    ensure_dirs()
    
    # Test verileri
    test_data = [
        {
            'timestamp': '2025-10-06 15:30:00',
            'sicaklik': 450.5,
            'titresim': 2.1,
            'tork': 65.3,
            'rul': 15.2,
            'status': 'CRITICAL'
        },
        {
            'timestamp': '2025-10-06 15:31:00',
            'sicaklik': 380.2,
            'titresim': 1.8,
            'tork': 45.7,
            'rul': 35.8,
            'status': 'PLANNED'
        },
        {
            'timestamp': '2025-10-06 15:32:00',
            'sicaklik': 320.1,
            'titresim': 1.2,
            'tork': 30.5,
            'rul': 75.4,
            'status': 'NORMAL'
        }
    ]
    
    # Test verilerini logla
    for data in test_data:
        append_prediction_log(data)
    
    print(f"✅ {len(test_data)} test verisi loglandı")
    
    # Günlük rapor oluştur
    try:
        excel_file = daily_report_to_excel()
        print(f"✅ Test raporu oluşturuldu: {excel_file}")
        return True
    except (OSError, ValueError) as e:
        print(f"❌ Test başarısız: {e}")
        return False


if __name__ == '__main__':
    # Test sistemini çalıştır
    success = test_reporting_system()
    
    if success:
        print("\n🎉 Raporlama sistemi başarıyla test edildi!")
    else:
        print("\n💥 Test başarısız!")
